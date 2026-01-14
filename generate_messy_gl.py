"""
Generate a large, messy, but balanced General Ledger Excel file for testing.

This script creates a realistic GL file with:
- 5,000-10,000+ transactions
- Messy formatting (inconsistent dates, mixed formats)
- Balanced debits and credits
- Headers, totals, subtotals
- Parent/subaccount structure
"""

import pandas as pd
import numpy as np
import random
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Set random seed for reproducibility
random.seed(42)
np.random.seed(42)

# Configuration
NUM_TRANSACTIONS = 8000  # Target number of transactions
START_DATE = datetime(2023, 1, 1)
END_DATE = datetime(2024, 12, 31)
OUTPUT_FILE = "sample_data/messy_large_gl_test.xlsx"

# Account structure with hierarchy
ACCOUNTS = {
    "Assets": {
        "Current Assets": ["Cash", "Accounts Receivable", "Inventory", "Prepaid Expenses"],
        "Fixed Assets": ["Equipment", "Vehicles", "Buildings", "Accumulated Depreciation"],
    },
    "Liabilities": {
        "Current Liabilities": ["Accounts Payable", "Accrued Expenses", "Short-term Loans"],
        "Long-term Liabilities": ["Long-term Loans", "Notes Payable"],
    },
    "Equity": {
        "Equity": ["Owner's Equity", "Retained Earnings", "Common Stock"],
    },
    "Revenue": {
        "Revenue": ["Sales Revenue", "Service Revenue", "Product Sales", "Consulting Fees", 
                   "Subscription Revenue", "Interest Income"],
    },
    "COGS": {
        "COGS": ["Cost of Goods Sold", "Direct Materials", "Direct Labor", "Manufacturing Overhead"],
    },
    "OpEx": {
        "Operating Expenses": ["Salaries Expense", "Rent Expense", "Utilities Expense", 
                               "Office Supplies", "Marketing Expense", "Travel Expense", 
                               "Insurance Expense", "Legal Expense", "Professional Fees"],
    },
    "Interest": {
        "Interest": ["Interest Expense", "Bank Fees"],
    },
    "Taxes": {
        "Taxes": ["Income Tax Expense", "Payroll Tax Expense", "Sales Tax"],
    },
    "D&A": {
        "D&A": ["Depreciation Expense", "Amortization Expense"],
    },
}

# Date format variations
DATE_FORMATS = [
    "%Y-%m-%d",      # 2024-01-15
    "%m/%d/%Y",      # 01/15/2024
    "%d-%b-%Y",      # 15-Jan-2024
    "%Y/%m/%d",      # 2024/01/15
    "%m-%d-%Y",      # 01-15-2024
]

# Column name variations
COLUMN_NAMES = {
    "date": ["Date", "Transaction Date", "Txn Date", "Posting Date", "DATE"],
    "account": ["Account", "Account Name", "Account Description", "ACCOUNT", "Account #"],
    "description": ["Description", "Memo", "Notes", "Transaction Description", "DESC"],
    "debit": ["Debit", "Debit Amount", "DR", "Debits", "DEBIT"],
    "credit": ["Credit", "Credit Amount", "CR", "Credits", "CREDIT"],
}

def generate_date_string(date_obj):
    """Generate date in random format (sometimes as text, sometimes as date)"""
    # NOTE: This function is NOT used for transaction rows anymore
    # Transaction rows always get valid dates
    # This is only for headers/totals which should have invalid dates
    format_type = random.choice(DATE_FORMATS)
    if random.random() < 0.3:  # 30% chance as text string
        return date_obj.strftime(format_type)
    return date_obj  # Return as date object

def generate_amount_string(amount):
    """Generate amount in random format"""
    # Only 1% chance of invalid text (to ensure proper parsing)
    if random.random() < 0.01:  # 1% chance of text
        return random.choice(["N/A", "Pending"])
    
    if random.random() < 0.15:  # 15% chance with currency symbol (as text)
        return f"${amount:,.2f}"
    elif random.random() < 0.25:  # 25% chance as text number
        return f"{amount:,.2f}"
    # 60% chance as actual number (best for parsing)
    return round(amount, 2)  # Return as number, rounded to 2 decimals

def flatten_accounts(accounts_dict):
    """Flatten account hierarchy into list of (parent, sub, account) tuples"""
    flat = []
    for parent, subs in accounts_dict.items():
        for sub, accounts in subs.items():
            for account in accounts:
                flat.append((parent, sub, account))
    return flat

def generate_transactions():
    """Generate balanced transactions"""
    all_accounts = flatten_accounts(ACCOUNTS)
    transactions = []
    
    # Track running balance
    total_debits = 0.0
    total_credits = 0.0
    
    # Generate date range
    date_range = (END_DATE - START_DATE).days
    
    for i in range(NUM_TRANSACTIONS):
        # Generate random date
        days_offset = random.randint(0, date_range)
        txn_date = START_DATE + timedelta(days=days_offset)
        
        # Determine transaction type
        txn_type = random.choice(["revenue", "expense", "transfer", "balance_sheet"])
        
        if txn_type == "revenue":
            # Revenue transaction: Credit revenue, debit asset
            revenue_account = random.choice([a for a in all_accounts if "Revenue" in a[0]])
            asset_account = random.choice([a for a in all_accounts if "Assets" in a[0] and "Cash" in a[2]])
            amount = random.uniform(100, 100000)
            
            # Debit asset
            transactions.append({
                "date": txn_date,
                "account_parent": asset_account[0],
                "account_sub": asset_account[1],
                "account": asset_account[2],
                "description": f"Payment received - {revenue_account[2]}",
                "debit": amount,
                "credit": 0,
            })
            total_debits += amount
            
            # Credit revenue
            transactions.append({
                "date": txn_date,
                "account_parent": revenue_account[0],
                "account_sub": revenue_account[1],
                "account": revenue_account[2],
                "description": f"Sales - {revenue_account[2]}",
                "debit": 0,
                "credit": amount,
            })
            total_credits += amount
            
        elif txn_type == "expense":
            # Expense transaction: Debit expense, credit asset/liability
            expense_account = random.choice([a for a in all_accounts if any(x in a[0] for x in ["OpEx", "COGS", "Interest", "Taxes", "D&A"])])
            payment_account = random.choice([a for a in all_accounts if "Assets" in a[0] and "Cash" in a[2]])
            amount = random.uniform(50, 50000)
            
            # Debit expense
            transactions.append({
                "date": txn_date,
                "account_parent": expense_account[0],
                "account_sub": expense_account[1],
                "account": expense_account[2],
                "description": f"Payment - {expense_account[2]}",
                "debit": amount,
                "credit": 0,
            })
            total_debits += amount
            
            # Credit payment account
            transactions.append({
                "date": txn_date,
                "account_parent": payment_account[0],
                "account_sub": payment_account[1],
                "account": payment_account[2],
                "description": f"Payment made - {expense_account[2]}",
                "debit": 0,
                "credit": amount,
            })
            total_credits += amount
            
        elif txn_type == "transfer":
            # Transfer: Debit one account, credit another (same type)
            account1 = random.choice(all_accounts)
            account2 = random.choice([a for a in all_accounts if a[0] == account1[0] and a != account1])
            amount = random.uniform(100, 100000)
            
            transactions.append({
                "date": txn_date,
                "account_parent": account1[0],
                "account_sub": account1[1],
                "account": account1[2],
                "description": f"Transfer to {account2[2]}",
                "debit": amount,
                "credit": 0,
            })
            total_debits += amount
            
            transactions.append({
                "date": txn_date,
                "account_parent": account2[0],
                "account_sub": account2[1],
                "account": account2[2],
                "description": f"Transfer from {account1[2]}",
                "debit": 0,
                "credit": amount,
            })
            total_credits += amount
            
        else:  # balance_sheet
            # Balance sheet transaction
            asset_account = random.choice([a for a in all_accounts if "Assets" in a[0]])
            liability_account = random.choice([a for a in all_accounts if "Liabilities" in a[0]])
            amount = random.uniform(1000, 200000)
            
            transactions.append({
                "date": txn_date,
                "account_parent": asset_account[0],
                "account_sub": asset_account[1],
                "account": asset_account[2],
                "description": f"Balance sheet entry",
                "debit": amount,
                "credit": 0,
            })
            total_debits += amount
            
            transactions.append({
                "date": txn_date,
                "account_parent": liability_account[0],
                "account_sub": liability_account[1],
                "account": liability_account[2],
                "description": f"Balance sheet entry",
                "debit": 0,
                "credit": amount,
            })
            total_credits += amount
    
    # Add balancing entry if needed
    difference = total_debits - total_credits
    if abs(difference) > 0.01:
        # Add small balancing entry
        balancing_account = random.choice([a for a in all_accounts if "Equity" in a[0]])
        if difference > 0:
            transactions.append({
                "date": END_DATE,
                "account_parent": balancing_account[0],
                "account_sub": balancing_account[1],
                "account": balancing_account[2],
                "description": "Balancing entry",
                "debit": 0,
                "credit": abs(difference),
            })
            total_credits += abs(difference)
        else:
            transactions.append({
                "date": END_DATE,
                "account_parent": balancing_account[0],
                "account_sub": balancing_account[1],
                "account": balancing_account[2],
                "description": "Balancing entry",
                "debit": abs(difference),
                "credit": 0,
            })
            total_debits += abs(difference)
    
    print(f"Generated {len(transactions)} transactions")
    print(f"Total Debits: ${total_debits:,.2f}")
    print(f"Total Credits: ${total_credits:,.2f}")
    print(f"Difference: ${abs(total_debits - total_credits):,.2f}")
    
    return transactions, total_debits, total_credits

def create_messy_dataframe(transactions):
    """Create DataFrame with messy formatting"""
    rows = []
    
    # Add header row (sometimes)
    if random.random() < 0.5:
        rows.append({
            "Date": "Date",
            "Account": "Account",
            "Description": "Description",
            "Debit": "Debit",
            "Credit": "Credit",
        })
    
    # Group transactions by account parent for structure
    current_parent = None
    
    for txn in transactions:
        # Sometimes add parent account header (these will have invalid dates - that's OK)
        # But reduce frequency to keep invalid dates < 10%
        if current_parent != txn["account_parent"] and random.random() < 0.08:  # Reduced to 8%
            rows.append({
                "Date": "",  # Empty date (will be filtered as invalid, which is correct for headers)
                "Account": txn["account_parent"],  # Parent header
                "Description": "",
                "Debit": "",
                "Credit": "",
            })
            current_parent = txn["account_parent"]
        
        # Build account name (sometimes flat, sometimes hierarchical)
        if random.random() < 0.4:  # 40% hierarchical
            account_name = f"{txn['account_parent']} : {txn['account_sub']} : {txn['account']}"
        elif random.random() < 0.3:  # 30% with sub
            account_name = f"{txn['account_sub']} : {txn['account']}"
        else:  # 30% flat
            account_name = txn['account']
        
        # Add messy formatting
        account_name = account_name.strip()
        if random.random() < 0.1:  # 10% with extra spaces
            account_name = "  " + account_name + "  "
        if random.random() < 0.15:  # 15% mixed case
            account_name = ''.join(c.upper() if random.random() < 0.3 else c.lower() for c in account_name)
        
        # CRITICAL: Transaction rows MUST have valid dates (so they're not filtered out)
        # Only use invalid dates for headers/totals, not transactions
        # Always use valid date for transactions - format as string sometimes, but still parseable
        if random.random() < 0.4:  # 40% as formatted string (still valid/parseable)
            format_type = random.choice(DATE_FORMATS)
            date_value = txn["date"].strftime(format_type)
        else:
            # 60% as date object (best for parsing)
            date_value = txn["date"]
        
        # Generate messy amounts (but keep as numbers mostly for balance)
        # Use actual numbers for transaction rows to ensure balance
        if txn["debit"] > 0:
            # 80% as number, 20% as formatted string (reduced text to maintain balance)
            debit_value = generate_amount_string(txn["debit"]) if random.random() < 0.2 else round(txn["debit"], 2)
        else:
            debit_value = ""
            
        if txn["credit"] > 0:
            # 80% as number, 20% as formatted string
            credit_value = generate_amount_string(txn["credit"]) if random.random() < 0.2 else round(txn["credit"], 2)
        else:
            credit_value = ""
        
        rows.append({
            "Date": date_value,
            "Account": account_name,
            "Description": txn["description"],
            "Debit": debit_value,
            "Credit": credit_value,
        })
        
        # Sometimes add empty row (very rare - these have invalid dates)
        if random.random() < 0.003:  # 0.3% chance (very rare)
            rows.append({
                "Date": "",
                "Account": "",
                "Description": "",
                "Debit": "",
                "Credit": "",
            })
    
    # Add total rows at the end (these will be filtered out, but show the balance)
    # Calculate from actual transaction amounts (not from formatted strings)
    total_debits = sum(t["debit"] for t in transactions)
    total_credits = sum(t["credit"] for t in transactions)
    
    rows.append({
        "Date": "",  # Empty date so it gets filtered
        "Account": "TOTAL",
        "Description": "",
        "Debit": total_debits,
        "Credit": total_credits,
    })
    
    return pd.DataFrame(rows)

def main():
    """Main function to generate messy GL file"""
    print("Generating large messy GL file...")
    print(f"Target transactions: {NUM_TRANSACTIONS}")
    
    # Generate transactions
    transactions, total_debits, total_credits = generate_transactions()
    
    # Create DataFrame
    df = create_messy_dataframe(transactions)
    
    print(f"\nDataFrame created with {len(df)} rows")
    print(f"Columns: {list(df.columns)}")
    
    # Ensure output directory exists
    output_path = Path(OUTPUT_FILE)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Save to Excel with messy formatting
    print(f"\nSaving to {OUTPUT_FILE}...")
    
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='General Ledger', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['General Ledger']
        
        # Apply messy formatting
        for row in range(2, min(len(df) + 2, 1000)):  # Format first 1000 rows
            for col in range(1, 6):
                cell = worksheet.cell(row, col)
                
                # Random formatting
                if random.random() < 0.1:  # 10% bold
                    cell.font = Font(bold=True)
                if random.random() < 0.05:  # 5% different font size
                    cell.font = Font(size=random.choice([9, 10, 11, 12, 14]))
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 50
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 15
    
    print(f"[OK] File saved: {OUTPUT_FILE}")
    print(f"\nFile Statistics:")
    print(f"  - Total rows: {len(df):,}")
    print(f"  - Total debits: ${total_debits:,.2f}")
    print(f"  - Total credits: ${total_credits:,.2f}")
    print(f"  - Difference: ${abs(total_debits - total_credits):,.2f}")
    print(f"\n[OK] Ready for testing in QoE Tool!")

if __name__ == "__main__":
    main()

